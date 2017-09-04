#!/home/mikepartain/GIT/python/bin/python
import os, sys, pprint
from openpyxl import *
os.system('clear')

wb = load_workbook('workbook.xlsx')
ws_list = {}
def select_sheet():
    ws_check = 'False'
    ws_range=[]
    print 'Please select the worksheet you would like to work with:'
    for i in range(len(wb.get_sheet_names())):
        ws_count = i+1
        if ws_count not in ws_range:
            ws_range.append(ws_count)

    # This section enumerates the worksheets and assigns a number to each sheet
    # and displays a list of the sheets and the number.  At this point the user
    # can select the worksheet they are interested in.
    for (ws_num, ws) in zip(range(len(wb.get_sheet_names())), wb.get_sheet_names()):
        print ws_num+1, ws
        ws_list[ws_num+1] = ws

    selected_ws = raw_input('Select Sheet number: ')
    for wsnum, wsname in ws_list.items():
        if int(selected_ws) == wsnum:
            ws_check = 'True'
            get_row_from_sheet(wsname)

    if ws_check == 'False':
        print 'Error %s, not a valid selection. Please try again.' % (selected_ws)
        select_sheet()
    
def get_row_from_sheet(wsname):
    global ws_range
    ws_range=[]
    ws=wb.get_sheet_by_name(wsname)
    spoke_row = input('%s has %s rows.  Please select the row for the spoke you are configuring.' % (wsname,ws.max_row))
    spoke_row = int(spoke_row)

    hostname = ws.cell(row=spoke_row, column=1).value
    description = ws.cell(row=spoke_row, column=2).value
    loopback = ws.cell(row=spoke_row, column=3).value
    aor = ws.cell(row=spoke_row, column=4).value
    pri_tunnel = ws.cell(row=spoke_row, column=5).value
    vlan = ws.cell(row=spoke_row, column=6).value
    value7 = ws.cell(row=spoke_row, column=7).value
    value8 = ws.cell(row=spoke_row, column=8).value

    print '%s, %s, %s %s, %s, %s, %s, %s' % ( hostname, description, loopback, aor, pri_tunnel, vlan, value7, value8)


select_sheet()

