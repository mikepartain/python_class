#!/home/mikepartain/GIT/python/bin/python
import os, sys, pprint
from openpyxl import *
os.system('clear')

wb = load_workbook('workbook.xlsx')
ws_list = {}
def select_sheet():
    ws_check = 'False'
    ws_range=[]
    print 'Please select the worksheet you would like to work with:'
    for i in range(len(wb.get_sheet_names())):
        ws_count = i+1
        if ws_count not in ws_range:
            ws_range.append(ws_count)

    # This section enumerates the worksheets and assigns a number to each sheet
    # and displays a list of the sheets and the number.  At this point the user
    # can select the worksheet they are interested in.
    for (ws_num, ws) in zip(range(len(wb.get_sheet_names())), wb.get_sheet_names()):
        print ws_num+1, ws
        ws_list[ws_num+1] = ws

    selected_ws = raw_input('Select Sheet number: ')
    for wsnum, wsname in ws_list.items():
        if int(selected_ws) == wsnum:
            ws_check = 'True'
            get_rows_from_sheet(wsname)

    if ws_check == 'False':
        print 'Error %s, not a valid selection. Please try again.' % (selected_ws)
        select_sheet()
    
def get_rows_from_sheet(wsname):
    global ws_range
    rows_data = []
    ws_range=[]
    ws=wb.get_sheet_by_name(wsname)
    print '%s has %s rows.' % (wsname,ws.max_row)
    for i in range(ws.max_row):
        i = i+1
        if i not in ws_range:
            ws_range.append(i)


    # for i in range(ws.max_row):
    print len(ws_range)
    print ws.max_row


    for cell in ws.iter_rows(min_row=2, max_col=4):
        print cell[0].value.ljust(20), cell[1].value.ljust(10),cell[2].value.ljust(30),cell[3].value



select_sheet()

