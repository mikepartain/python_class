#!/home/mikepartain/GIT/python/bin/python
import os, sys, pprint, ipaddr, string, subprocess
from openpyxl import *
from netaddr import *

os.system('clear')

wb = load_workbook('dmvpn.xlsx')
ws_list = {}
device = {}

def select_sheet():
    ws_check = 'False'
    ws_range=[]
    print 'Please select the worksheet you would like to work with:'
    for i in range(len(wb.get_sheet_names())):
        ws_count = i+1
        if ws_count not in ws_range:
            ws_range.append(ws_count)

    # This section enumerates the worksheets and assigns a number to each sheet
    # and displays a list of the sheets and the number.  At this point the user
    # can select the worksheet they are interested in.
    for (ws_num, ws) in zip(range(len(wb.get_sheet_names())), wb.get_sheet_names()):
        print ws_num+1, ws
        ws_list[ws_num+1] = ws

    selected_ws = raw_input('Select Sheet number: ')
    for wsnum, wsname in ws_list.items():
        if int(selected_ws) == wsnum:
            ws_check = 'True'
            get_row_from_sheet(wsname)

    if ws_check == 'False':
        print 'Error %s, not a valid selection. Please try again.' % (selected_ws)
        select_sheet()
    
def get_row_from_sheet(wsname):
    global ws_range, aor, hostname, description, loopback, pri_tunnel, pri_tunnel_ip, sec_tunnel, sec_tunnel_ip, vlan

    ws_range=[]
    ws=wb.get_sheet_by_name(wsname)
    spoke_row = input('%s has %s rows.  Please select the row for the spoke you are configuring.' % (wsname,ws.max_row))
    spoke_row = int(spoke_row)

    aor = wsname
    hostname = ws.cell(row=spoke_row, column=1).value
    description = ws.cell(row=spoke_row, column=2).value
    loopback = ws.cell(row=spoke_row, column=3).value
    pri_tunnel = ws.cell(row=spoke_row, column=4).value
    pri_tunnel_ip = ws.cell(row=spoke_row, column=5).value
    sec_tunnel = ws.cell(row=spoke_row, column=6).value
    sec_tunnel_ip = ws.cell(row=spoke_row, column=7).value
    vlan = ws.cell(row=spoke_row, column=8).value

    device.update(HOSTNAME=hostname)
    device.update(DESCRIPTION=description)
    device.update(LOOPBACK=loopback)


    get_router(hostname,aor,pri_tunnel)

def get_router(HOSTNAME,AOR,pri_tunnel):
    print ("""2. Select Router Model for %s:
        A. Cisco x800 (1800, 2800, 3800)
        B. Cisco x800 W/Switch Module
        C. Cisco x900 (1900, 2900, 3900)
        D. Cisco x900 W/Switch Module
        E. Cisco  800 Series (819)

    """) % HOSTNAME
    router = raw_input("2. Select Router Type: ")
    router = router.upper()
    if router == 'A':
        ext_intereface = "FastEthernet0/0"
        inside_int = 'FastEthernet0/1'
        device.update(OUTSIDE_INT=ext_intereface, INSIDE_INT=inside_int, MODEL='x800')

    elif router == 'B':
        ext_intereface = "FastEthernet0/0"
        inside_int = 'FastEthernet0/1'
        device.update(OUTSIDE_INT=ext_intereface, INSIDE_INT=inside_int, MODEL='x800', SWITCH='SWITCH')


    elif router == 'C':
        ext_intereface = "GigabitEthernet0/0"
        inside_int = 'GigabitEthernet0/1'
        device.update(OUTSIDE_INT=ext_intereface, INSIDE_INT=inside_int, MODEL='x900')

    elif router == 'D':
        ext_intereface = "GigabitEthernet0/0"
        inside_int = 'GigabitEthernet0/1'
        device.update(OUTSIDE_INT=ext_intereface, INSIDE_INT=inside_int, MODEL='x900', SWITCH='SWITCH')

    elif router == 'E':
        ext_intereface = "FastEthernet0/0"
        inside_int = 'FastEthernet0/1'
        device.update(OUTSIDE_INT=ext_intereface, INSIDE_INT=inside_int, MODEL='800', SWITCH='SWITCH')

    else:
        print "Come on... Simple instructions.  You can do this!  Lets try again."
        return get_router(HOSTNAME)

    get_community(HOSTNAME,AOR,pri_tunnel)

def get_community(HOSTNAME,community,pri_tunnel):
    print 'Community is: %s' % (community)
    pri_tunnel = int(pri_tunnel)
    # while True:
        # print ("""
        # Select the Community for %s.
        # A. WBN (1569)
        # B. SGT (1562)
        # C. HFL (1566)
        # ZZ. EXIT""") % HOSTNAME
        #
        # community = raw_input("3. Select Community: ")
        # community = community.upper()
        #print type(community)
    if community == 'WBN':
        AOR = 'WBN'
        AS = '1569'
        device.update(AOR=AOR, AS=AS, PRIMARY='TLAW', SECONDARY='tlas', P_NHRP_MAP='139.139.3.35', S_NHRP_MAP='139.139.19.35')
        tunnel(AOR, AS)

    elif community == 'SGT':
        AOR = 'SGT'
        AS = '1562'
        device.update(AOR=AOR, AS=AS, PRIMARY='TLAS', SECONDARY='tlaw', P_NHRP_MAP='139.139.19.35', S_NHRP_MAP='139.139.3.35')
        tunnel(AOR,AS)


    elif community == 'HFL':
        AOR = 'HFL'
        AS='1566'
        device.update(AOR=AOR, AS=AS, PRIMARY='TLAS', SECONDARY='tlaw', P_NHRP_MAP='139.139.19.35', S_NHRP_MAP='139.139.3.35')
        tunnel(AOR,AS)

    elif community == 'KLN':
        AOR = 'KLN'
        AS = '1580'
        device.update(AOR=AOR, AS=AS, PRIMARY='TLAW', SECONDARY='tlas', P_NHRP_MAP='139.139.3.35', S_NHRP_MAP='139.139.19.35')
        tunnel(AOR, AS)

    elif community == 'ZZ':
        print "Goodbye."
        sys.exit()

    elif community != "":
        print "Not a valid choice."

def tunnel(AOR, AS):
    while True:
        tunnel_num = str(pri_tunnel)
        tunnel_num = tunnel_num[4:]
        tunnel_num = int(tunnel_num)
        # as_tunnel = AS
        as_tunnel = int(AS)
        as_tunnel = as_tunnel * 100
        # print tunnel_num[4:]
        # tunnel_num = pri_tunnel[4:]
        # print 'Tunnel number is %s%s' % (AS,tunnel_num)
        # tunnel_num = int(raw_input('4. Enter the tunnel number this will be for %s: in AS %s.' % (AOR, AS)))
        if tunnel_num > 0 and tunnel_num < 80:
            # print "This is a dual homed tunnel."
            secondary = as_tunnel + tunnel_num + 1
            #print "Secondary:", secondary
            device.update(CONN='DUAL', SEC_TUNNEL=secondary)
            AS2 = int(AS) * 100 + int(tunnel_num)
            tunnel_num = as_tunnel + tunnel_num
            device.update(PRI_TUNNEL=tunnel_num)

        elif tunnel_num > 1800:
            #print "This is a single homed tunnel"
            device.update(CONN='SINGLE')
            AS2 = int(AS)*10000 + int(tunnel_num)
            tunnel_num = AS2
            device.update(PRI_TUNNEL=tunnel_num)


        elif tunnel_num < 1 or tunnel_num > 99:
            print "Enter a correct tunnel number between 1 and 99"
            return tunnel(AOR, AS)

        get_tunnel_ip(AOR, tunnel_num)
        break

def get_tunnel_ip(AOR, pri_tunnel):
    #print device
    tunnel_mask = ''
    if 'SEC_TUNNEL' in device:
        #print "Processing DUAL Homed spoke."
        if 'TLAW' in device.values():
            #print "Processing TLAW Primary Spoke."
            # tunnel_ip = raw_input("5. Enter the Primary tunnel IP Address and CIDR (139.139.160.0/20): ")
            tunnel_ip = pri_tunnel_ip
            tunnel_ip = ipaddr.IPv4Network(tunnel_ip)
            pri_nhrp = tunnel_ip[1]
            sec_nhrp = tunnel_ip[1] + 4352
            secondary_tunnel_ip = tunnel_ip.ip + 4352
            secondary_tunnel_ip = ipaddr.IPv4Network(secondary_tunnel_ip)
            #print "New secondary IP:", secondary_tunnel_ip
            get_user_ip()
            device.update(SEC_TUNNEL_IP=str(secondary_tunnel_ip.ip))
            device.update(PRI_TUNNEL_IP=str(tunnel_ip.ip), TUNNEL_MASK=str(tunnel_ip.netmask))
            device.update(PRI_NHRP=pri_nhrp, SEC_NHRP=sec_nhrp)

        elif 'TLAS' in device.values():
            #print "Processing TLAS Primary Spoke."
            tunnel_ip = raw_input("5. Enter the Primary tunnel IP Address and CIDR (139.139.176.0/20): ")
            tunnel_ip = ipaddr.IPv4Network(tunnel_ip)
            pri_nhrp = tunnel_ip[1]
            sec_nhrp = tunnel_ip[1] - 3840
            secondary_tunnel_ip = tunnel_ip.ip - 3840
            secondary_tunnel_ip = ipaddr.IPv4Network(secondary_tunnel_ip)
            #print "New secondary IP:", secondary_tunnel_ip
            get_user_ip()
            device.update(SEC_TUNNEL_IP=str(secondary_tunnel_ip.ip))
            device.update(PRI_TUNNEL_IP=str(tunnel_ip.ip), TUNNEL_MASK=str(tunnel_ip.netmask))
            device.update(PRI_NHRP=pri_nhrp, SEC_NHRP=sec_nhrp)




    elif 'SEC_TUNNEL' not in device:
        print "Processing single homed spoke."
        tunnel_ip = raw_input("5. Enter the tunnel IP Address and CIDR (2.2.2.2/27): ")
        tunnel_ip = ipaddr.IPAddress(tunnel_ip)
        device.update(PRI_TUNNEL_IP=str(tunnel_ip.ip), TUNNEL_MASK=str(tunnel_ip.netmask))
        get_user_ip()


    else:
        print "Fatal error..."
        sys.exit()

def get_user_ip():
    # user_network = raw_input("7. Enter a router inside IP address with CIDR (3.3.3.3/28): ")
    user_subnet = ipaddr.IPv4Network(vlan)


    # if user_subnet.is_private or user_subnet.is_multicast or user_subnet.is_reserved:
    #     print "Enter a valid IP Network address."
    #     return get_user_ip()


    user_mask = IPNetwork(vlan)
    voip = str(user_mask.ip)
    voip = voip.split('.',-1)[1:4]
    voip = '.'.join(voip)
    voip = '10.'+voip
    voip_pool = str(user_mask.network)
    voip_pool = voip_pool.split('.',-1)[1:4]
    voip_pool = '.'.join(voip_pool)
    voip_pool = '10.'+str(voip_pool)
    voip_int = voip.split('.',-1)[1:5]
    voip_int = '.'.join(voip_int)
    voip_int = '10.'+str(voip_int)
    device.update(INSIDE_IP=str(user_subnet.ip), INSIDE_MASK=str(user_subnet.netmask), INSIDE_NET=str(user_mask.network), VOIP=str(voip), VOIP_POOL=str(voip_pool), VOIP_INT=str(voip_int))

def display_config():
    if 'SEC_TUNNEL' in device and 'SWITCH' not in device:
        template = open('roles/dmvpn/vars/main.tpl')
        src = string.Template(template.read())
        HOSTNAME=device['HOSTNAME']
        DESCRIPTION=device['DESCRIPTION']
        AOR=device['AOR']
        AS=device['AS']
        CONN=device['CONN']
        LOOPBACK=device['LOOPBACK']
        PRIMARY=device['PRIMARY']
        SECONDARY=device['SECONDARY']
        OUTSIDE_INT=device['OUTSIDE_INT']
        INSIDE_INT=device['INSIDE_INT']
        PRI_TUNNEL=device['PRI_TUNNEL']
        PRI_TUNNEL_IP=device['PRI_TUNNEL_IP']
        PRI_NHRP=device['PRI_NHRP']
        P_NHRP_MAP=device['P_NHRP_MAP']
        TUNNEL_MASK=device['TUNNEL_MASK']
        SEC_TUNNEL =device['SEC_TUNNEL']
        SEC_TUNNEL_IP=device['SEC_TUNNEL_IP']
        SEC_NHRP = device['SEC_NHRP']
        S_NHRP_MAP = device['S_NHRP_MAP']
        INSIDE_IP=device['INSIDE_IP']
        INSIDE_MASK=device['INSIDE_MASK']
        INSIDE_NET=device['INSIDE_NET']
        VOIP=device['VOIP']
        VOIP_POOL=device['VOIP_POOL']


        os.system('clear')
        print """Below is your configuration items:
        Configuration:
        Hostname:               %s
        Description:            %s
        AOR:                    %s
        AS:                     %s
        MODE:                   %s
        LOOPBACK IP:            %s
        PRIMARY SITE:           %s
        SECONDARY SITE:         %s
        EXTERNAL INTERFACE:     %s
        INSIDE   INTERFACE:     %s
        PRIMARY TUNNEL:         %s
        PRIMARY TUNNEL IP:      %s
        PRIMARY TUNNEL MASK:    %s
        SECONDARY TUNNEL:       %s
        SECONDARY TUNNEL IP:    %s
        SECONDARY TUNNEL MASK:  %s
        INSIDE INT IP ADDR:     %s
        INSIDE INT NETMASK:     %s
        DHCP EXCLUDE:           %s
        DHCP POOL:              %s %s
        VOIP EXCLUDE:           %s %s
        VOIP POOL:              %s
        PRI_NHRP:               %s
        SEC_NHRP:               %s
        P_NHRP_MAP:             %s
        S_NHRP_MAP:             %s


        """ % (HOSTNAME, DESCRIPTION, AOR, AS, CONN, LOOPBACK, PRIMARY, SECONDARY, OUTSIDE_INT, INSIDE_INT, PRI_TUNNEL, PRI_TUNNEL_IP, TUNNEL_MASK, SEC_TUNNEL, SEC_TUNNEL_IP, TUNNEL_MASK, INSIDE_IP, INSIDE_MASK, INSIDE_IP, INSIDE_NET, INSIDE_MASK, VOIP, INSIDE_MASK, VOIP_POOL, PRI_NHRP, SEC_NHRP, P_NHRP_MAP, S_NHRP_MAP)

        d = {'HOSTNAME': HOSTNAME, 'DESCRIPTION': DESCRIPTION, 'AOR': AOR, 'AS': AS, 'MODE': CONN, 'PRIMARY_SITE': PRIMARY, 'LOOPBACK': LOOPBACK, 'EXTERNAL_INT': OUTSIDE_INT, 'INSIDE_INT': INSIDE_INT,
             'PRI_TUNNEL': PRI_TUNNEL, 'SECONDARY_SITE': SECONDARY, 'PRI_NHRP': PRI_NHRP, 'SEC_NHRP': SEC_NHRP, 'P_NHRP_MAP': P_NHRP_MAP, 'S_NHRP_MAP': S_NHRP_MAP,
             'PRI_TUNNEL_IP': PRI_TUNNEL_IP, 'TUNNEL_MASK': TUNNEL_MASK, 'INSIDE_IP': INSIDE_IP, 'INSIDE_MASK': INSIDE_MASK, 'INSIDE_NET': INSIDE_NET,
             'SEC_TUNNEL': SEC_TUNNEL, 'SEC_TUNNEL_IP': SEC_TUNNEL_IP, 'VOIP_IP': VOIP, 'VOIP_POOL': VOIP_POOL}

        result = src.substitute(d)
        fileout = open('roles/dmvpn/vars/main.yml', 'w')
        fileout.write(result)

    elif 'SEC_TUNNEL' in device and 'SWITCH' in device:
        template = open('roles/dmvpn/vars/main_sw.tpl')
        src = string.Template(template.read())
        HOSTNAME = device['HOSTNAME']
        DESCRIPTION = device['DESCRIPTION']
        AOR = device['AOR']
        AS = device['AS']
        CONN = device['CONN']
        LOOPBACK = device['LOOPBACK']
        PRIMARY = device['PRIMARY']
        SECONDARY = device['SECONDARY']
        OUTSIDE_INT = device['OUTSIDE_INT']
        INSIDE_INT = device['INSIDE_INT']
        PRI_TUNNEL = device['PRI_TUNNEL']
        PRI_TUNNEL_IP = device['PRI_TUNNEL_IP']
        PRI_NHRP = device['PRI_NHRP']
        P_NHRP_MAP = device['P_NHRP_MAP']
        TUNNEL_MASK = device['TUNNEL_MASK']
        SEC_TUNNEL = device['SEC_TUNNEL']
        SEC_TUNNEL_IP = device['SEC_TUNNEL_IP']
        SEC_NHRP = device['SEC_NHRP']
        S_NHRP_MAP = device['S_NHRP_MAP']
        INSIDE_IP = device['INSIDE_IP']
        INSIDE_MASK = device['INSIDE_MASK']
        INSIDE_NET = device['INSIDE_NET']
        VOIP = device['VOIP']
        VOIP_POOL = device['VOIP_POOL']
        SWITCH = device['SWITCH']

        os.system('clear')
        print """Below is your configuration items:
            Configuration:
            Hostname:               %s
            Description:            %s
            AOR:                    %s
            AS:                     %s
            MODE:                   %s
            LOOPBACK IP:            %s
            PRIMARY SITE:           %s
            SECONDARY SITE:         %s
            EXTERNAL INTERFACE:     %s
            INSIDE   INTERFACE:     %s
            PRIMARY TUNNEL:         %s
            PRIMARY TUNNEL IP:      %s
            PRIMARY TUNNEL MASK:    %s
            SECONDARY TUNNEL:       %s
            SECONDARY TUNNEL IP:    %s
            SECONDARY TUNNEL MASK:  %s
            INSIDE INT IP ADDR:     %s
            INSIDE INT NETMASK:     %s
            DHCP EXCLUDE:           %s
            DHCP POOL:              %s %s
            VOIP EXCLUDE:           %s %s
            VOIP POOL:              %s
            SWITCH:                 %s
            PRI_NHRP:               %s
            SEC_NHRP:               %s
            P_NHRP_MAP:             %s
            S_NHRP_MAP:             %s


            """ % (HOSTNAME, DESCRIPTION, AOR, AS, CONN, LOOPBACK, PRIMARY, SECONDARY, OUTSIDE_INT, INSIDE_INT, PRI_TUNNEL, PRI_TUNNEL_IP, TUNNEL_MASK, SEC_TUNNEL, SEC_TUNNEL_IP, TUNNEL_MASK, INSIDE_IP, INSIDE_MASK, INSIDE_IP, INSIDE_NET, INSIDE_MASK, VOIP, INSIDE_MASK, VOIP_POOL, SWITCH, PRI_NHRP, SEC_NHRP, P_NHRP_MAP, S_NHRP_MAP)

        d = {'HOSTNAME': HOSTNAME, 'DESCRIPTION': DESCRIPTION, 'AOR': AOR, 'AS': AS, 'MODE': CONN,
             'PRIMARY_SITE': PRIMARY, 'LOOPBACK': LOOPBACK, 'EXTERNAL_INT': OUTSIDE_INT, 'INSIDE_INT': INSIDE_INT,
             'PRI_TUNNEL': PRI_TUNNEL, 'SECONDARY_SITE': SECONDARY, 'PRI_NHRP': PRI_NHRP, 'SEC_NHRP': SEC_NHRP, 'P_NHRP_MAP': P_NHRP_MAP, 'S_NHRP_MAP': S_NHRP_MAP,
             'PRI_TUNNEL_IP': PRI_TUNNEL_IP, 'TUNNEL_MASK': TUNNEL_MASK, 'INSIDE_IP': INSIDE_IP,
             'INSIDE_MASK': INSIDE_MASK, 'INSIDE_NET': INSIDE_NET,
             'SEC_TUNNEL': SEC_TUNNEL, 'SEC_TUNNEL_IP': SEC_TUNNEL_IP, 'VOIP_IP': VOIP, 'VOIP_POOL': VOIP_POOL,
             'SWITCH': SWITCH}


        result = src.substitute(d)
        fileout = open('roles/dmvpn/vars/main.yml', 'w')
        fileout.write(result)

        print 'You new DMVPN Config with switch module is located in the vars folder called main.yml'

    elif 'SEC_TUNNEL' not in device:
        HOSTNAME = device['HOSTNAME']
        AOR = device['AOR']
        AS = device['AS']
        CONN = device['CONN']
        LOOPBACK = device['LOOPBACK']
        PRIMARY = device['PRIMARY']
        OUTSIDE_INT = device['OUTSIDE_INT']
        INSIDE_INT = device['INSIDE_INT']
        PRI_TUNNEL = device['PRI_TUNNEL']
        PRI_TUNNEL_IP = device['PRI_TUNNEL_IP']
        MASK = device['MASK']
        INSIDE_IP = device['INSIDE_IP']
        INSIDE_MASK = device['INSIDE_MASK']
        INSIDE_NET = device['INSIDE_NET']

        print """Below is your configuration items:
                    Configuration:
                    Hostname:               %s
                    AOR:                    %s
                    AS:                     %s
                    MODE:                   %s
                    LOOPBACK IP:            %s
                    EXTERNAL INTERFACE:     %s
                    INSIDE   INTERFACE      %s
                    PRI_TUNNEL:             %s
                    PRI_TUNNEL IP:          %s
                    PRI_TUNNEL MASK:        %s
                    INSIDE INT IP ADDR:     %s
                    INSIDE INT NETMASK:     %s
                    DHCP EXCLUDE:           %s
                    DHCP POOL:              %s %s %s %s


                    """ % (HOSTNAME, AOR, AS, CONN, LOOPBACK, PRIMARY, OUTSIDE_INT, INSIDE_INT, PRI_TUNNEL, PRI_TUNNEL_IP, MASK, INSIDE_IP, INSIDE_MASK, INSIDE_NET, INSIDE_IP, INSIDE_NET, INSIDE_MASK)

        # print device
        # single_commit(device['HOSTNAME'], device['LOOPBACK'], device['INT'], device['INSIDE_INT'], device['PRI_TUNNEL'], device['IP'], device['MASK'], device['INSIDE_IP'], device['INSIDE_MASK'], device['INSIDE_NET'])

def gen_config():
	subprocess.call('ansible-playbook site.yml', shell=True)

def deploy_config():
    from napalm import get_network_driver
    driver = get_network_driver('ios')
    print 'The configuration has been generated.  '
    deploy = raw_input('Would you like to deploy this config now? (Y/N)')
    deploy = deploy.upper().strip()
    if deploy == 'Y':
        os.system('clear')
        print '''To successfully deploy this config, ensure the following are met:
        Router has the below configuration:
            A. External interface (Gig 0/0 or Fas0/0) is connected
            B. IP address is set
                a.  Configured for DHCP and has been obtained
                b.  Static IP is configured and reachable
            C. Cisco Local user account is created
            D. SSH is enabled
                a. Set Hostname
                b. Set domain name
                c. Generate crypto key

        Example Config:
            config t
            hostname VPN-SPOKE
            ip domain name lab.com
            !
            interface g0/0
            ! interface f0/0
            ! interface e0/0
            ip address dhcp
            no shut
            !
            username cisco priv 15 password cisco
            !
            line con 0
             privi level 15
            line vty 0 4
             transport input ssh
             login local
             privi level 15
            !
            crypto key gen rsa mod 2048


        '''

        HOST = raw_input('Please enter the IP of the target router. ')
        CONFIG = 'CFGS/'+hostname+'.config'

        device = driver(HOST, 'cisco', 'cisco')
        print 'Connecting to %s' % hostname
        device.open()

        print 'Loading Candidate Config to device.'
        device.load_replace_candidate(filename=CONFIG)

        print 'Loading Merge Config to device.'
        device.load_merge_candidate(filename=CONFIG)

        print 'Comparing configurations.'
        diffs = device.compare_config()
        #print diffs


        print '\n\nConfiguration has been copied to %s.\n  ' % hostname

        commit = raw_input('Would you like to commit this configuration to %s? (Y/N)' % hostname)
        commit = commit.upper().strip()
        if commit == 'Y':
            device.commit_config()
            print 'Configuration has been deployed to the running config.  \n' \
                  'Login to your device and validate the configuration.\n' \
                  'If the configuration is what you expect, dont forget the \n' \
                  'write mem or copy run start'

            # rollback = raw_input('Would you like to rollback the configuration? (Y/N)')
            # rollback = rollback.upper().strip()
            # if rollback == 'Y':
            #     print 'Please wait...  Attempting to rollback deployed config.'
            #     device.rollback()
            # else:
            #     print 'Please login to your device and perform a copy run start!'

        else:
            print "Thanks for using Mike's DMVPN Creator."
            print "Your config has been saved to CFGS/%s." % hostname


    else:
        print "Your config has been saved to CFGS/%s." % hostname
        sys.exit()

    device.close()




select_sheet()
display_config()
gen_config()
deploy_config()

print device
